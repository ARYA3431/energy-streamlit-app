import streamlit as st
import pandas as pd
import datetime
import os
from openpyxl import load_workbook

# ==============================
# FILE SETTINGS
# ==============================

FILE_NAME = "Energy Sheet.xlsx"

if not os.path.exists(FILE_NAME):
    st.error("Excel file not found!")
    st.stop()

current_month = datetime.datetime.now().strftime("%B")
today_str = datetime.datetime.now().strftime("%d-%m-%Y")

st.title("⚡ Energy Monitoring System")

# ==============================
# HELPER FUNCTIONS
# ==============================

def clean_text(text):
    return str(text).upper().replace("-", "").replace("#", "").replace(" ", "")

def update_excel(ws, col_index, name, value):
    clean_name = clean_text(name)

    for row in range(4, ws.max_row + 1):
        col1 = ws.cell(row=row, column=1).value
        col2 = ws.cell(row=row, column=2).value

        combined = f"{col1} {col2}"
        clean_combined = clean_text(combined)

        if clean_name in clean_combined:
            ws.cell(row=row, column=col_index).value = int(value)
            return

def get_previous_total(ws, col_index, name):

    prev_col = col_index - 1

    if prev_col < 3:
        return 0

    for row in range(4, ws.max_row + 1):
        col1 = str(ws.cell(row=row, column=1).value)
        col2 = str(ws.cell(row=row, column=2).value)

        combined = (col1 + " " + col2).upper()

        if name.upper() in combined:
            value = ws.cell(row=row, column=prev_col).value

            if value is None:
                return 0

            try:
                return float(value)
            except:
                return 0

    return 0

# ==============================
# INPUT GRID
# ==============================

def input_grid(labels):
    values = {}
    for i in range(0, len(labels), 3):
        cols = st.columns(3)
        for j, label in enumerate(labels[i:i+3]):
            with cols[j]:
                values[label] = st.number_input(label, step=1.0, key=label)
    return values

# ==============================
# LABELS
# ==============================

tr_labels = [
    "TR-1 (31.5 MVA)", "TR-2 (31.5 MVA)", "TR-3 (31.5 MVA)",
    "TR-4 (31.5 MVA)", "TR-5 (31.5 MVA)"
]

lhf_labels = ["LHF#1", "LHF#2"]

lcss9_labels = ["LCSS-9 FDR-1", "LCSS-9 FDR-3", "LCSS-9 FDR-2"]
lcss8_labels = ["LCSS-8 FDR-1", "LCSS-8 FDR-3", "LCSS-8 FDR-2"]

ccm_labels = ["CCM-1 EMS-1", "CCM-1 EMS-2"]

fan_labels = [
    "PRIMARY ID FAN #1", "PRIMARY ID FAN #2",
    "SECONDARY ID FAN#1", "SECONDARY ID FAN#2", "SECONDARY ID FAN#3"
]

rcph_labels = ["RCPH I/C-1", "RCPH I/C-2"]

lcp_labels = ["LCP FDR-1", "LCP FDR-3"]

other_labels = ["Grinder I/C Caster"]

heat_labels = ["No. of Heat Tap", "No. of Heat Cast"]

# ==============================
# INPUT SECTION
# ==============================

st.header("Enter Meter Readings")

tr_values = input_grid(tr_labels)
lhf_values = input_grid(lhf_labels)
lcss9_values = input_grid(lcss9_labels)
lcss8_values = input_grid(lcss8_labels)
ccm_values = input_grid(ccm_labels)
fan_values = input_grid(fan_labels)
rcph_values = input_grid(rcph_labels)
lcp_values = input_grid(lcp_labels)
other_values = input_grid(other_labels)
heat_values = input_grid(heat_labels)

# ==============================
# LIVE CALCULATION
# ==============================
st.subheader("⚡ Live Calculation") 
total_tr = sum(tr_values.values()) 
total_lf = sum(lhf_values.values()) 
total_lcp = sum(lcp_values.values()) 
total_caster = ( sum(lcss8_values.values()) + sum(lcss9_values.values()) + sum(ccm_values.values()) + other_values["Grinder I/C Caster"] ) 
total_bof = total_tr - total_lcp - total_caster 
total_rcph = sum(rcph_values.values())
heat_tap = heat_values["No. of Heat Tap"]
heat_cast = heat_values["No. of Heat Cast"]


# ==============================
# DISPLAY
# ==============================

col1, col2, col3 = st.columns(3)
col1.metric("TOTAL TR", int(total_tr))
col2.metric("TOTAL LF", int(total_lf))
col3.metric("TOTAL LCP", int(total_lcp))

col4, col5, col6 = st.columns(3)
col4.metric("TOTAL CASTER", int(total_caster))
col5.metric("TOTAL BOF", int(total_bof))
col6.metric("TOTAL RCPH", int(total_rcph))

col7, col8 = st.columns(2)
col7.metric("HEAT TAP", int(heat_tap))
col8.metric("HEAT CAST", int(heat_cast))



# ==============================
# SUBMIT
# ==============================

if st.button("Submit"):

    wb = load_workbook(FILE_NAME, data_only=False)
    ws = wb[current_month]

    # Find column
    col_index = None
    for col in range(3, ws.max_column + 1):
        if str(ws.cell(row=2, column=col).value) == today_str:
            col_index = col
            break

    if col_index is None:
        col_index = ws.max_column + 1
        ws.cell(row=2, column=col_index).value = today_str

    # ==============================
    # PREVIOUS VALUES
    # ==============================

    prev_lcp = get_previous_total(ws, col_index, "TOTAL LCP CONSUMPTION")
    prev_rcph = get_previous_total(ws, col_index, "TOTAL RCPH CONSUMPTION")
    prev_caster = get_previous_total(ws, col_index, "TOTAL CASTER CONSUMPTION")
    prev_bof = get_previous_total(ws, col_index, "TOTAL BOF CONSUMPTION")
    prev_lf = get_previous_total(ws, col_index, "TOTAL LF CONSUMPTION")
    prev_total = get_previous_total(ws, col_index, "Total")
    prev_pid1 = get_previous_total(ws, col_index, "PRIMARY ID FAN #1")
    prev_pid2 = get_previous_total(ws, col_index, "PRIMARY ID FAN #2")
    prev_sid1 = get_previous_total(ws, col_index, "SECONDARY ID FAN#1")
    prev_sid2 = get_previous_total(ws, col_index, "SECONDARY ID FAN#2")
    prev_sid3 = get_previous_total(ws, col_index, "SECONDARY ID FAN#3")

    # ==============================
    # PER DAY
    # ==============================

    lcp_per_day = total_lcp - prev_lcp
    rcph_per_day = total_rcph - prev_rcph
    caster_per_day = total_caster - prev_caster + rcph_per_day*0.65
    bof_per_day = total_bof - prev_bof - 0.65*rcph_per_day - lcp_per_day
    lf_per_day = total_lf - prev_lf
    total_energy_per_day = caster_per_day + bof_per_day + lf_per_day
    pid1_per_day = fan_values["PRIMARY ID FAN #1"] - prev_pid1
    pid2_per_day = fan_values["PRIMARY ID FAN #2"] - prev_pid2
    sid1_per_day = fan_values["SECONDARY ID FAN#1"] - prev_sid1
    sid2_per_day = fan_values["SECONDARY ID FAN#2"] - prev_sid2
    sid3_per_day = fan_values["SECONDARY ID FAN#3"] - prev_sid3
    total_ID_Fan = pid1_per_day + pid2_per_day + sid1_per_day + sid2_per_day + sid3_per_day
    caster_per_ton = caster_per_day/(heat_cast*347)
    bof_per_ton = bof_per_day/(heat_tap*347)
    lf_per_ton = lf_per_day/(heat_tap*347)
    total_per_ton = caster_per_ton + bof_per_ton + lf_per_ton
    
     
    

    # ==============================
    # SAVE INPUTS
    # ==============================

    for group in [
        tr_values, lhf_values, lcss9_values, lcss8_values,
        ccm_values, fan_values, rcph_values, lcp_values, other_values
    ]:
        for key, val in group.items():
            update_excel(ws, col_index, key, val)

    update_excel(ws, col_index, "No. of Heat Tap", heat_tap)
    update_excel(ws, col_index, "No. of Heat Cast", heat_cast)

    # ==============================
    # SAVE TOTALS
    # ==============================

    update_excel(ws, col_index, "Total", total_tr)
    update_excel(ws, col_index, "TOTAL LF CONSUMPTION", total_lf)
    update_excel(ws, col_index, "TOTAL LCP CONSUMPTION", total_lcp)
    update_excel(ws, col_index, "TOTAL CASTER CONSUMPTION", total_caster)
    update_excel(ws, col_index, "TOTAL BOF CONSUMPTION", total_bof)
    update_excel(ws, col_index, "TOTAL RCPH CONSUMPTION", total_rcph)

    # ==============================
    # SAVE PER DAY
    # ==============================

    update_excel(ws, col_index, "LCP CONSUMPTION PER DAY", lcp_per_day)
    update_excel(ws, col_index, "RCPH CONSUMPTION PER DAY", rcph_per_day)
    update_excel(ws, col_index, "CASTER CONSUMPTION PER DAY", caster_per_day)
    update_excel(ws, col_index, "BOF CONSUMPTION PER DAY", bof_per_day)
    update_excel(ws, col_index, "LF CONSUMPTION PER DAY", lf_per_day)
    update_excel(ws, col_index, "TOTAL ENERGY CONSUMPTION PER DAY", total_energy_per_day)
    update_excel(ws, col_index, "CONSUMPTION PRIMARY ID FAN #1", pid1_per_day)
    update_excel(ws, col_index, "CONSUMPTION PRIMARY ID FAN #2", pid2_per_day)
    update_excel(ws, col_index, "CONSUMPTION SECONDARY ID FAN #1", sid1_per_day)
    update_excel(ws, col_index, "CONSUMPTION SECONDARY ID FAN #2", sid2_per_day)
    update_excel(ws, col_index, "CONSUMPTION SECONDARY ID FAN #3", sid3_per_day)
    update_excel(ws, col_index, "TOTAL ID FAN CONSUMPTION", total_ID_Fan)
    update_excel(ws, col_index, "CONSUMPTION PER TON (CASTER)", caster_per_ton)
    update_excel(ws, col_index, "CONSUMPTION PER TON (BOF)", bof_per_ton)
    update_excel(ws, col_index, "CONSUMPTION PER TON (LF)", lf_per_ton)
    update_excel(ws, col_index, "TOTAL CONSUMPTION PER TON", total_per_ton)

    # SAVE FILE
    wb.calculation.fullCalcOnLoad = True
    wb.save(FILE_NAME)
    st.session_state["caster_per_day"] = caster_per_day
    st.session_state["bof_per_day"] = bof_per_day
    st.session_state["lf_per_day"] = lf_per_day
    st.session_state["total_energy_per_day"] = total_energy_per_day
    
    st.session_state["caster_per_ton"] = caster_per_ton
    st.session_state["bof_per_ton"] = bof_per_ton
    st.session_state["lf_per_ton"] = lf_per_ton
    st.session_state["total_per_ton"] = total_per_ton

    st.success("✅ Data Saved Successfully")

# ==============================
# DISPLAY TABLE
# ==============================

wb_data = load_workbook(FILE_NAME, data_only=True)
ws_data = wb_data[current_month]

data = list(ws_data.values)
df = pd.DataFrame(data[2:], columns=data[1])

new_cols = list(df.columns[:2])
for col in df.columns[2:]:
    try:
        new_cols.append(pd.to_datetime(col).strftime("%d-%m-%Y"))
    except:
        new_cols.append(col)

df.columns = new_cols

for col in df.columns[2:]:
    df[col] = pd.to_numeric(df[col], errors='coerce')

df = df.fillna("")

for col in df.columns[2:]:
    df[col] = df[col].apply(lambda x: int(x) if x != "" else "")


required_keys = [
    "caster_per_day",
    "bof_per_day",
    "lf_per_day",
    "total_energy_per_day",
    "caster_per_ton",
    "bof_per_ton",
    "lf_per_ton",
    "total_per_ton"
]

if all(key in st.session_state for key in required_keys):

    st.subheader("📊 Final Calculated Values")

    col1, col2, col3 = st.columns(3)

    col1.metric("CASTER PER DAY", int(st.session_state["caster_per_day"]))
    col2.metric("BOF PER DAY", int(st.session_state["bof_per_day"]))
    col3.metric("LF PER DAY", int(st.session_state["lf_per_day"]))

    col4, col5, col6 = st.columns(3)

    col4.metric("TOTAL PER DAY", round(st.session_state["total_energy_per_day"]))
    col5.metric("CASTER PER TON", round(st.session_state["caster_per_ton"], 2))
    col6.metric("BOF PER TON", round(st.session_state["bof_per_ton"], 2))

    col7, col8 = st.columns(2)

    col7.metric("LF PER TON", round(st.session_state["lf_per_ton"], 2))
    col8.metric("TOTAL PER TON", round(st.session_state["total_per_ton"], 2))

st.subheader("📊 Energy Data")
st.dataframe(df, use_container_width=True)

# DOWNLOAD
with open(FILE_NAME, "rb") as file:
    st.download_button("📥 Download Excel", file, "Energy Sheet.xlsx")
